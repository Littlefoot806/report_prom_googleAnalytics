import analytics_selenium
import Prom_selenium 
from get_data_from_xls import get_data_from_xls
from openpyxl import workbook
from datetime import datetime, date, timedelta
import argparse
import logging
import sys


logging.basicConfig(filename='output.log',level=logging.INFO, format='%(asctime)s, %(levelname)s: %(message)s in -> %(module)s, line is %(lineno)d', datefmt='%m/%d/%Y %I:%M:%S %p')

parser = argparse.ArgumentParser()
parser.add_argument('day', type=int, help='How much day in past')
args = parser.parse_args()
day = int(args.day)

wb = workbook.Workbook()
ws = wb.active

try:
	logging.info('Started parsing xls')
	data_from_xls = get_data_from_xls('17.xls')
except Exception as e:
	logging.critical(str(e))
	sys.exit()

logging.info('Have a result ')

yesterday = date.today() - timedelta(day)
try:
	logging.info('Writing to xlsx')
	ws['A2'] = str(yesterday.strftime('%d.%m.%Y'))
	file_name = str(datetime.today().strftime('%d-%m-%y, %H,%M,%S'))+'.xlsx'
	if data_from_xls:
		ws['A1'] = 'General statistics'
		for k, v in data_from_xls.items():
			ws["{0}2".format(k)] = v
except Exception as e:
	logging.critical(str(e))
	sys.exit()
logging.info('Finished parsing xls')
print('Finished parsing xls')

try:
	logging.info('Started Prom parsing')
	data_from_prom = Prom_selenium.main(day)
except Exception as e:
	logging.critical(str(e))
	sys.exit()
logging.info('Have a result ')
try:
	logging.info('Writing to xlsx')
	if data_from_prom[0]['company_id'] == '2128639':
		ws["A3"] = 'prosale: 33 Korovy'
		for k, v in data_from_prom[0].items():
			if k == 'company_id':
				continue
			if k == 'orders':
				i = 9
				ws['A8'] = '33 Korovy'
				for key, val in v.items():
					ws['A{}'.format(i)] = 'Prosale'
					ws['B{}'.format(i)] = key
					val = val.replace('+38', '')
					val = val[:3]+'-'+val[3:6]+'-'+val[6:8]+'-'+val[8:]
					ws['C{}'.format(i)] = val
					i += 1
				continue
			ws["{}4".format(k)] = v
	if data_from_prom[1]['company_id'] == '2361594':
		ws["A5"] = 'prosale: krolikam'
		for k, v in data_from_prom[1].items():
			if k == 'company_id':
				continue
			if k == 'orders':
				i = 9
				ws['E8'] = 'krolikam'
				for key, val in v.items():
					ws['E{}'.format(i)] = 'Prosale'
					ws['F{}'.format(i)] = key
					val = val.replace('+38', '')
					val = val[:3]+'-'+val[3:6]+'-'+val[6:8]+'-'+val[8:]
					ws['G{}'.format(i)] = val
					i += 1
				continue
			ws["{}6".format(k)] = v
except Exception as e:
	logging.critical(str(e))
	sys.exit()
logging.info('Finished Prom parsing')

try:
	logging.info('Started Analytics parsing')
	data_from_analytics = analytics_selenium.main(day)
except Exception as e:
	logging.critical(str(e))
	sys.exit()
logging.info('Have a result ')

try:
	for data_analytics in data_from_analytics:
		if data_analytics['acc_id'] == 'a59793658w94063945p97997794':
			ws['n2'] = data_analytics['sessions']
			ws['o2'] = data_analytics['social']
		elif data_analytics['acc_id'] == 'a89214817w132397505p136356129':
			ws['t2'] = data_analytics['sessions']
			ws['u2'] = data_analytics['social']
		elif data_analytics['acc_id'] == 'a34451481w62001262p63538862':
			ws['h2'] = data_analytics['sessions']
			ws['i2'] = data_analytics['social']
		elif data_analytics['acc_id'] == 'a69426406w106425247p110797471':
			ws['b2'] = data_analytics['sessions']
			ws['c2'] = data_analytics['social']
		elif data_analytics['acc_id'] == 'a75713134w114181655p119300061':
			ws['z2'] = data_analytics['sessions']
			ws['aa2'] = data_analytics['social']
except Exception as e:
	logging.critical(str(e))
	sys.exit()
logging.info('Finished, file is {}'.format(file_name))

wb.save(filename=file_name)
	
